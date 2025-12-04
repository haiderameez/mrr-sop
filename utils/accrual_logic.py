import re
import logging
from datetime import datetime
from typing import TYPE_CHECKING

import pandas as pd
import numpy as np
from openpyxl import load_workbook

from .functions import (
    extract_all, normalize_name, get_month_start_end, months_list_from_field,
    overlap_days, safe_round
)

if TYPE_CHECKING:
    from app import WorkflowConfig

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

MONTH_NAME_MAP = {
    month: datetime(2000, month, 1).strftime('%b')
    for month in range(1, 13)
}

def process_accrual_update(config: "WorkflowConfig"):
    invoice_path = config.invoice_file
    master_path = config.master_file
    mis_path = config.mis_path
    master_sheet = config.master_sheet_name
    mis_sheet = config.mis_sheet_name
    target_month = config.target_month
    target_year = config.target_year
    month_name = MONTH_NAME_MAP[target_month]

    logger.info("Loading invoice data from %s", invoice_path)
    invoice = pd.read_excel(invoice_path)
    logger.info("Invoice rows loaded: %d", len(invoice))
    invoice = invoice[~invoice["Invoice ID"].duplicated(keep=False)]
    logger.info("Invoice rows after deduplication: %d", len(invoice))
    invoice = invoice[["Invoice Number", "Customer Name"]]

    logger.info("Loading master sheet %s from %s", master_sheet, master_path)
    master = pd.read_excel(master_path, skiprows=2, sheet_name=master_sheet)
    master.columns = master.columns.str.strip()
    master = master.drop(master.columns[10:], axis=True)
    master = master.rename(columns={"Invoice": "Invoice Number"})

    # Normalize Master for matching
    master_norm = master.copy()
    master_norm['All Invoices'] = master_norm['Invoice Number'].apply(extract_all)
    master_norm = master_norm.explode('All Invoices')

    merged = invoice.merge(
        master_norm,
        left_on='Invoice Number',
        right_on='All Invoices',
        how='left'
    )
    logger.info("Merged invoice/master rows: %d", len(merged))

    logger.info("Loading MIS sheet %s from %s", mis_sheet, mis_path)
    mis = pd.read_excel(mis_path, skiprows=2, sheet_name=mis_sheet)
    mis = mis.rename(columns=lambda x: x.strip() if isinstance(x, str) else x)

    if 'Invoice' in mis.columns and 'Invoice Number' not in mis.columns:
        mis = mis.rename(columns={"Invoice": "Invoice Number"})

    month_start, month_end = get_month_start_end(target_month, target_year)

    # Pre-calculations for MIS
    mis['Customer Name'] = mis['Customer Name'].fillna('')
    mis['cn_norm'] = mis['Customer Name'].apply(normalize_name)

    logger.info("Beginning MIS enrichment across %d rows", len(mis))
    for idx, row in mis.iterrows():
        start = pd.to_datetime(row.get('Start Date'), errors='coerce')
        end = pd.to_datetime(row.get('End Date'), errors='coerce')
        if pd.notna(start) and pd.notna(end):
            mis.at[idx, 'Contract period'] = (end - start).days + 1

    # Prepare merged data for lookup
    merged = merged.copy()
    if 'Customer Name_y' not in merged.columns and 'Customer Name' in merged.columns:
        merged = merged.rename(columns={'Customer Name': 'Customer Name_y'})
    merged = merged.rename(columns=lambda x: x.strip() if isinstance(x, str) else x)
    merged['Customer Name_y'] = merged['Customer Name_y'].fillna('')
    merged['cn_norm'] = merged['Customer Name_y'].apply(normalize_name)

    merged_map = {}
    for _, r in merged.iterrows():
        merged_map.setdefault(r['cn_norm'], []).append(r)

    
    col_name_days = f'Days in {month_name}'
    col_name_sales_days = f'{month_name} Sales in days'
    col_name_sales_months = f'{month_name} Sales in months'

    for idx, row in mis.iterrows():
        norm = row.get('cn_norm', '')
        start = pd.to_datetime(row.get('Start Date'), errors='coerce')
        end = pd.to_datetime(row.get('End Date'), errors='coerce')
        months_present = months_list_from_field(row.get('Months'), target_year)
        overlap = pd.notna(start) and pd.notna(end) and not (end < month_start or start > month_end)
        applicable = (target_month in months_present) or overlap
        
        if norm in merged_map:
            mrow = merged_map[norm][-1]
            inv_new = mrow.get('Invoice Number_y') or mrow.get('Invoice') or mrow.get('Invoice Number')
            inv_old = row.get('Invoice Number')
            
            if pd.isna(inv_old) or str(inv_old).strip() == "":
                if pd.notna(inv_new):
                    mis.at[idx, 'Invoice Number'] = inv_new
            else:
                vals = [x.strip() for x in re.split(r'[;,]', str(inv_old)) if x.strip()]
                for p in re.split(r'[;,]', str(inv_new) if pd.notna(inv_new) else ''):
                    p = p.strip()
                    if p and p not in vals:
                        vals.append(p)
                mis.at[idx, 'Invoice Number'] = ", ".join(vals) if vals else inv_old
            
            pc_new = mrow.get('Payment Cycle')
            pc_old = row.get('Payment Cycle')
            if (pd.isna(pc_old) or str(pc_old).strip() == "") and pd.notna(pc_new):
                mis.at[idx, 'Payment Cycle'] = pc_new
            
            if pd.isna(start):
                start = pd.to_datetime(mrow.get('Start Date'), errors='coerce')
                if pd.notna(start):
                    mis.at[idx, 'Start Date'] = start
            
            if pd.isna(end):
                end = pd.to_datetime(mrow.get('End Date'), errors='coerce')
                if pd.notna(end):
                    mis.at[idx, 'End Date'] = end
        
        if applicable and pd.notna(start) and pd.notna(end):
            contract_period_days = (end - start).days + 1
            mis.at[idx, 'Contract period'] = contract_period_days
            d_month = overlap_days(start, end, month_start, month_end)
            amt = row.get('Contract Amount') if pd.notna(row.get('Contract Amount')) else 0
            sales_days = (amt / contract_period_days) * d_month if contract_period_days > 0 else 0
            
            months_diff = (end.year - start.year) * 12 + (end.month - start.month)
            if end.day >= start.day:
                mc = max(1, months_diff + 1)
            else:
                mc = max(1, months_diff)
            monthwise = amt / mc if mc > 0 else 0
            
            closing = (end.year == target_year and end.month == target_month)
            
            if target_month in months_present:
                sales_months = 0 if closing else monthwise
            elif overlap:
                sales_months = monthwise
            else:
                sales_months = np.nan
            
            mis.at[idx, col_name_days] = d_month
            mis.at[idx, col_name_sales_days] = sales_days
            if pd.notna(sales_months):
                mis.at[idx, col_name_sales_months] = sales_months

    existing = set(mis['cn_norm'].fillna(''))
    to_add = []

    for _, r in merged.iterrows():
        nm = r['cn_norm']
        if nm in existing or nm == "":
            continue
        
        s = pd.to_datetime(r.get('Start Date'), errors='coerce')
        e = pd.to_datetime(r.get('End Date'), errors='coerce')
        overlap = pd.notna(s) and pd.notna(e) and not (e < month_start or s > month_end)
        
        if not overlap:
            continue
        
        new = pd.Series(index=mis.columns, dtype=object)
        if 'Months' in mis.columns:
            new['Months'] = f'{month_name}-{str(target_year)[-2:]}'
        
        new['Customer Name'] = r.get('Customer Name_y')
        new['cn_norm'] = nm
        
        if 'Invoice Number' in mis.columns:
            inv_val = r.get('Invoice Number_y', r.get('Invoice Number', r.get('Invoice')))
            new['Invoice Number'] = inv_val
        
        if 'Nature of service' in mis.columns:
            new['Nature of service'] = r.get('Nature of service')
        
        new['Start Date'] = r.get('Start Date')
        new['End Date'] = r.get('End Date')
        new['Payment Cycle'] = r.get('Payment Cycle')
        new['Contract Amount'] = r.get('Contract Amount')
        
        if pd.notna(s) and pd.notna(e):
            contract_period_days = (e - s).days + 1
            new['Contract period'] = contract_period_days
            d_month = overlap_days(s, e, month_start, month_end)
            amt = new.get('Contract Amount') if pd.notna(new.get('Contract Amount')) else 0
            sales_days = (amt / contract_period_days) * d_month if contract_period_days > 0 else 0
            
            months_diff = (e.year - s.year) * 12 + (e.month - s.month)
            if e.day >= s.day:
                mc = max(1, months_diff + 1)
            else:
                mc = max(1, months_diff)
            monthwise = amt / mc if mc > 0 else 0
            
            closing = (e.year == target_year and e.month == target_month)
            
            if 'Months' in new.index:
                months_present_new = months_list_from_field(new.get('Months'), target_year)
            else:
                months_present_new = []
            
            if target_month in months_present_new:
                sales_months_new = 0 if closing else monthwise
            else:
                sales_months_new = monthwise
            
            new[col_name_days] = d_month
            new[col_name_sales_days] = sales_days
            new[col_name_sales_months] = sales_months_new
        
        to_add.append(new.to_dict())

    if to_add:
        new_df = pd.DataFrame(to_add)
        for c in mis.columns:
            if c not in new_df.columns:
                new_df[c] = np.nan
        new_df = new_df[mis.columns]
        mis = pd.concat([mis, new_df], ignore_index=True)

    # Rounding and Cleanup
    amount_columns = ['Contract Amount', col_name_sales_days, col_name_sales_months]
    for col in mis.columns:
        if any(keyword in str(col) for keyword in ['Sales', 'Invoiced amount', 'Days in']):
            if col not in amount_columns:
                amount_columns.append(col)

    for c in amount_columns:
        if c in mis.columns:
            mis[c] = mis[c].apply(safe_round)

    mis = mis.loc[:, ~mis.columns.duplicated()]
    if 'cn_norm' in mis.columns:
        mis = mis.drop(columns=['cn_norm'])

    mis = mis.rename(columns=lambda x: x.strip() if isinstance(x, str) else x)

    for col in [col_name_days, col_name_sales_days, col_name_sales_months]:
        if col not in mis.columns:
            mis[col] = np.nan

    # Writing back to Excel with formatting retention
    wb = load_workbook(mis_path)
    ws = wb[mis_sheet]

    header_row_excel = 3
    max_row = ws.max_row
    max_col = ws.max_column

    for col_idx, col_name in enumerate(mis.columns, start=1):
        ws.cell(row=header_row_excel, column=col_idx).value = col_name

    for row in range(header_row_excel + 1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            ws.cell(row=row, column=col_idx).value = None

    for r in range(mis.shape[0]):
        excel_row = header_row_excel + 1 + r
        for c_idx, col in enumerate(mis.columns, start=1):
            ws.cell(row=excel_row, column=c_idx).value = mis.iloc[r][col]

    logger.info("Writing updated MIS data back to %s", mis_path)
    wb.save(mis_path)
    logger.info("Accrual update complete")
    return mis  # Return DF for next steps
