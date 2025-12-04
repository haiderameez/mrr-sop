import logging
from typing import TYPE_CHECKING

import pandas as pd
from openpyxl import load_workbook

if TYPE_CHECKING:
    from app import WorkflowConfig

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def process_subscribers_and_churn(config: "WorkflowConfig"):
    mis_path = config.mis_path
    mis_sheet = config.mis_sheet_name
    logger.info("Loading addition sheet for subscriber tracking")
    add = pd.read_excel(mis_path, sheet_name='Addition')
    if add.columns[0].startswith("Unnamed") and "list of customer added" in str(add.iloc[0, 1]).lower():
        header_row = add[add.iloc[:, 1].astype(str).str.lower() == "s.no."].index[0]
        add = add.iloc[header_row:, 1:].reset_index(drop=True)

    logger.info("Loading deletions sheet for subscriber tracking")
    delete = pd.read_excel(mis_path, sheet_name='Deletions')
    if delete.columns[0].startswith("Unnamed") and "list of customer" in str(delete.iloc[0, 1]).lower():
        header_row = delete[delete.iloc[:, 1].astype(str).str.lower() == "s.no."].index[0]
        delete = delete.iloc[header_row:, 1:].reset_index(drop=True)

    # Identify Active Subscribers
    logger.info("Identifying active subscribers for churn calculation")
    mis = pd.read_excel(mis_path, skiprows=2, sheet_name=mis_sheet)
    mis = mis.rename(columns={"Invoice": "Invoice Number"})
    month_cols = [c for c in mis.columns if "Sales in months" in str(c)]
    last_month_col = month_cols[-1] if month_cols else None
    
    company_names = mis.loc[mis[last_month_col].notna(), 'Customer Name'].dropna().unique()
    active_sub = pd.read_excel(mis_path, sheet_name='Active Subscriber')
    
    month_name_cleaned = last_month_col.split()[0] # e.g., "Oct" from "Oct Sales in months"
    active_sub[month_name_cleaned] = pd.Series(company_names)

    # Determine Additions/Deletions
    cols = active_sub.columns[-2:]
    prev_col, last_col = cols[0], cols[1]
    prev_set = set(active_sub[prev_col].dropna())
    last_set = set(active_sub[last_col].dropna())

    new_customers = sorted(last_set - prev_set)
    removed_customers = sorted(prev_set - last_set)

    add[month_name_cleaned] = pd.Series([""] + list(new_customers)).reindex(add.index)
    delete[month_name_cleaned] = pd.Series([""] + list(removed_customers)).reindex(add.index)
    logger.info(
        "Calculated subscriber changes - additions: %d, deletions: %d",
        len(new_customers),
        len(removed_customers),
    )

    # --- Churnout Logic ---
    logger.info("Updating customer churnout sheet")
    churnout = pd.read_excel(mis_path, sheet_name='Customer churnout')
    
    # Cleanup Churnout format if needed
    second_col_name = churnout.columns[1]
    if second_col_name == 'Coresphere100 Technologies Pvt. Ltd.':
        churnout = pd.read_excel(mis_path, skiprows=4, sheet_name='Customer churnout')
        churnout = churnout.iloc[:, 1:]
        rows = {"begin": 0, "add": 1, "del": 2, "end": 3}
        cols_iter = list(churnout.columns[1:])
        for i, col in enumerate(cols_iter):
            if i == 0:
                begin = churnout.at[rows["begin"], col]
            else:
                begin = churnout.at[rows["end"], cols_iter[i - 1]]
                churnout.at[rows["begin"], col] = begin
            
            addi = churnout.at[rows["add"], col]
            dele = churnout.at[rows["del"], col]
            churnout.at[rows["end"], col] = begin + addi - dele

    last_end_value = churnout.iloc[3, -1]
    new_count = len(new_customers)
    del_count = len(removed_customers)
    new_month_end = last_end_value + new_count - del_count
    
    churnout[month_name_cleaned] = [last_end_value, new_count, del_count, new_month_end]

    # --- Final Excel Cleanup and Save ---
    logger.info("Removing outdated helper sheets from workbook")
    wb = load_workbook(mis_path)
    
    to_delete = [
        s for s in wb.sheetnames
        if "pivot" in s.lower() or s in ["Active Subscriber", "Addition", "Deletions", "MR-AR", "Customer churnout"]
    ]
    for s in to_delete:
        wb.remove(wb[s])
    wb.save(mis_path)

    logger.info("Writing subscriber-related sheets back to %s", mis_path)
    with pd.ExcelWriter(mis_path, engine="openpyxl", mode="a") as writer:
        active_sub.to_excel(writer, sheet_name="Active Subscriber", index=False)
        add.to_excel(writer, sheet_name='Addition', index=False)
        delete.to_excel(writer, sheet_name='Deletions', index=False)
        churnout.to_excel(writer, sheet_name="Customer churnout", index=False)
    logger.info("Subscriber processing complete")
    
