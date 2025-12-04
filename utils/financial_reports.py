import re
import logging
from typing import TYPE_CHECKING

import pandas as pd
import numpy as np
from openpyxl import load_workbook

from .functions import extract_all

if TYPE_CHECKING:
    from app import WorkflowConfig

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def generate_pivot_and_mrar(config: "WorkflowConfig"):
    mis_path = config.mis_path
    mis_sheet = config.mis_sheet_name
    logger.info("Reading invoices and accrual sheets from %s", mis_path)
    invoices = pd.read_excel(mis_path, sheet_name="Invoices", skiprows=2)
    fy_acc = pd.read_excel(mis_path, sheet_name=mis_sheet, skiprows=2)

    # --- B2B Logic ---
    b2b_df = invoices[invoices["Nature"].astype(str).str.upper() == "B2B"]
    b2b_companies = b2b_df[["Customer Name"]].drop_duplicates().reset_index(drop=True)
    b2b_data = fy_acc[fy_acc["Customer Name"].isin(b2b_companies["Customer Name"])].reset_index(drop=True)

    # 1. Pivot Month (Sales in months)
    fixed_month_cols = [
        "Invoiced amount for April", "Invoiced amount for MAY",
        "June in months", "July in months", "Aug Sales in months", "Sept Sales in months"
    ]
    all_dynamic = [
        c for c in b2b_data.columns
        if c not in fixed_month_cols and (
            c.endswith("Sales in months") or c.endswith("in months") or c.lower().startswith("invoiced amount for")
        )
    ]
    month_order_map = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
                       "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}
    
    def extract_month_index(col):
        txt = col.lower()
        for m in month_order_map:
            if m in txt:
                return month_order_map[m]
        return 999

    def month_sort_key(col):
        return (extract_month_index(col), col.lower())

    dynamic_month_cols = sorted(all_dynamic, key=extract_month_index)
    month_cols = fixed_month_cols + dynamic_month_cols

    logger.info("Building pivot_month for %d B2B rows", len(b2b_data))
    pivot_month = b2b_data.pivot_table(index="Payment Cycle", values=month_cols, aggfunc="sum").reset_index()
    value_month_cols = [c for c in month_cols if c in pivot_month.columns]
    value_month_cols = sorted(value_month_cols, key=month_sort_key)
    grand_total_pm = pivot_month[value_month_cols].sum().to_frame().T
    grand_total_pm.insert(0, "Payment Cycle", "Grand Total")
    pivot_month = pd.concat([pivot_month, grand_total_pm], ignore_index=True)
    pivot_month = pivot_month[["Payment Cycle"] + value_month_cols]

    # 2. Pivot Days (Sales in days)
    fixed_date_cols = [
        "April Sales (Days Wise)", "May Sales (Days Wise)",
        "June in days", "July in days", "Aug Sales in days", "Sept Sales in days"
    ]
    all_dynamic_days = [
        c for c in b2b_data.columns
        if c not in fixed_date_cols and (
            c.endswith("Sales in days") or c.endswith("in days") or "(days" in c.lower()
        )
    ]
    dynamic_date_cols = sorted(all_dynamic_days, key=extract_month_index)
    date_cols = fixed_date_cols + dynamic_date_cols

    logger.info("Building pivot_days for %d B2B rows", len(b2b_data))
    pivot_days = b2b_data.pivot_table(index="Payment Cycle", values=date_cols, aggfunc="sum").reset_index()
    value_day_cols = [c for c in date_cols if c in pivot_days.columns]
    value_day_cols = sorted(value_day_cols, key=month_sort_key)
    grand_total_pd = pivot_days[value_day_cols].sum().to_frame().T
    grand_total_pd.insert(0, "Payment Cycle", "Grand Total")
    pivot_days = pd.concat([pivot_days, grand_total_pd], ignore_index=True)
    pivot_days = pivot_days[["Payment Cycle"] + value_day_cols]

    # --- B2C Logic ---
    b2c_df = invoices[invoices["Nature"].astype(str).str.upper() == "B2C"].copy()
    b2c_df["Norm Inv"] = b2c_df["Invoice Number"].apply(extract_all)
    b2c_df = b2c_df.explode("Norm Inv")

    fy_acc["Norm Inv"] = fy_acc["Invoice Number"].apply(extract_all)
    fy_acc = fy_acc.explode("Norm Inv")
    b2c_data = fy_acc[fy_acc["Norm Inv"].isin(b2c_df["Norm Inv"])].reset_index(drop=True)

    # --- MR-AR Report Construction ---
    pivot = pivot_days.copy()
    fixed_day_map = {
        "April Sales (Days Wise)": "April", "May Sales (Days Wise)": "May",
        "June in days": "June", "July in days": "July",
        "Aug Sales in days": "August", "Sept Sales in days": "September"
    }
    available_fixed = {k: v for k, v in fixed_day_map.items() if k in pivot.columns}
    dynamic_day_cols = [c for c in pivot.columns if c not in available_fixed and ("days" in c.lower())]

    def extract_month_name(col):
        t = col.lower()
        for m in month_order_map:
            if m in t:
                return m.capitalize()
        return None

    dynamic_day_rename = {c: extract_month_name(c) for c in dynamic_day_cols if extract_month_name(c)}
    rename_map_all = {}
    rename_map_all.update(available_fixed)
    rename_map_all.update(dynamic_day_rename)

    pivot = pivot.rename(columns=rename_map_all)
    
    final_month_cols = [rename_map_all[c] for c in available_fixed] + sorted(
        [dynamic_day_rename[c] for c in dynamic_day_rename],
        key=lambda x: month_order_map[x[:3].lower()]
    )

    mr_ar_b2b = pivot[["Payment Cycle"] + final_month_cols].copy()
    mr_ar_b2b = mr_ar_b2b.rename(columns={"Payment Cycle": "Particulars"})

    # B2C Totals
    available_month_fixed = [c for c in fixed_month_cols if c in b2c_data.columns]
    dynamic_month_cols_b2c = [
        c for c in b2c_data.columns
        if c not in available_month_fixed and (
            c.endswith("Sales in months") or c.endswith("in months") or c.lower().startswith("invoiced amount for")
        )
    ]
    dynamic_month_cols_b2c = sorted(dynamic_month_cols_b2c, key=month_sort_key)
    dynamic_month_rename = {c: extract_month_name(c) for c in dynamic_month_cols_b2c if extract_month_name(c)}
    rename_map_months = {c: extract_month_name(c) for c in available_month_fixed}
    rename_map_months.update(dynamic_month_rename)

    all_month_cols_b2c = available_month_fixed + dynamic_month_cols_b2c
    b2c_totals = b2c_data[all_month_cols_b2c].sum().to_frame().T
    b2c_totals.insert(0, "Particulars", "B2C")
    b2c_totals = b2c_totals.rename(columns=rename_map_months)

    final_months = sorted(rename_map_months.values(), key=lambda x: month_order_map[x[:3].lower()])

    logger.info("Combining MR-AR dataframe with B2C totals")
    mr_ar = pd.concat([mr_ar_b2b, b2c_totals[["Particulars"] + final_months]], ignore_index=True)
    
    # Correction for B2C short names if present
    mr_ar.loc[mr_ar["Particulars"]=="B2C", ["April","June","July","August","September"]] = \
        mr_ar.loc[mr_ar["Particulars"]=="B2C", ["Apr","Jun","Jul","Aug","Sep"]].values
    mr_ar = mr_ar.drop(columns=["Apr", "Jun", "Jul", "Aug", "Sep"], errors="ignore")

    logger.info(
        "Generated MR-AR (%d rows), Pivot Days (%d rows), Pivot Month (%d rows)",
        len(mr_ar),
        len(pivot_days),
        len(pivot_month),
    )
    return mr_ar, pivot_days, pivot_month


def save_financial_reports(config: "WorkflowConfig", mr_ar_df, pivot_days, pivot_month):
    """Write MR-AR and pivot sheets back to the MIS workbook."""
    mis_path = config.mis_path
    logger.info("Saving financial reports to workbook: %s", mis_path)
    wb = load_workbook(mis_path)

    remove_targets = [
        sheet for sheet in wb.sheetnames
        if sheet.lower().startswith("pivot") or sheet in {"MR-AR"}
    ]
    for sheet_name in remove_targets:
        wb.remove(wb[sheet_name])
    wb.save(mis_path)

    with pd.ExcelWriter(mis_path, engine="openpyxl", mode="a") as writer:
        mr_ar_df.to_excel(writer, sheet_name="MR-AR", index=False)
        pivot_days.to_excel(writer, sheet_name="Pivot Days", index=False)
        pivot_month.to_excel(writer, sheet_name="Pivot Month", index=False)
    logger.info("Financial sheets updated successfully")
